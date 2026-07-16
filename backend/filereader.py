"""Convert uploaded files to HTML/text the model can read.

Supported: .docx (via mammoth), .xlsx (via openpyxl), .pdf (via pypdf),
.html passthrough, and any plain-text format. Extraction never raises —
failures come back as an error the model (and user) can see.
"""

import base64
import html
import io
from pathlib import Path

MAX_EXTRACT_CHARS = 20_000
MAX_TABLE_ROWS = 500


def _decode(data_url: str) -> bytes:
    b64 = data_url.split(",", 1)[1] if data_url.startswith("data:") else data_url
    return base64.b64decode(b64)


def _docx_to_html(raw: bytes) -> str:
    import mammoth

    return mammoth.convert_to_html(io.BytesIO(raw)).value


def _xlsx_to_html(raw: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = "".join(
                f"<td>{html.escape('' if c is None else str(c))}</td>"
                for c in row
            )
            rows.append(f"<tr>{cells}</tr>")
            if len(rows) >= MAX_TABLE_ROWS:
                rows.append("<tr><td>[... more rows truncated ...]</td></tr>")
                break
        parts.append(
            f"<h3>Sheet: {html.escape(ws.title)}</h3>"
            f"<table>{''.join(rows)}</table>"
        )
    wb.close()
    return "\n".join(parts)


def _pdf_to_html(raw: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    parts = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        parts.append(f"<h3>Page {i + 1}</h3>\n<pre>{html.escape(text)}</pre>")
        if sum(len(p) for p in parts) > MAX_EXTRACT_CHARS:
            break
    return "\n".join(parts)


def _text_to_html(raw: bytes) -> str:
    return f"<pre>{html.escape(raw.decode('utf-8', 'replace'))}</pre>"


def extract(name: str, data_url: str) -> dict:
    """Return {name, content (HTML), error}; content is '' on failure."""
    try:
        raw = _decode(data_url)
        ext = Path(name).suffix.lower()
        if ext == ".docx":
            content = _docx_to_html(raw)
        elif ext in (".xlsx", ".xlsm", ".xltx"):
            content = _xlsx_to_html(raw)
        elif ext == ".pdf":
            content = _pdf_to_html(raw)
        elif ext in (".html", ".htm"):
            content = raw.decode("utf-8", "replace")
        else:
            content = _text_to_html(raw)
        if len(content) > MAX_EXTRACT_CHARS:
            content = content[:MAX_EXTRACT_CHARS] + "\n[... truncated ...]"
        return {"name": name, "content": content, "error": None}
    except Exception as exc:
        return {"name": name, "content": "", "error": str(exc)[:200]}
